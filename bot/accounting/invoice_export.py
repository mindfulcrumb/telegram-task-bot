"""Export service for invoice data."""

from __future__ import annotations

import csv
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from bot.accounting.invoice_models import Invoice
from bot.accounting.categorizer import get_category_display
from bot.accounting.export_service import HEADER_FILL, HEADER_FONT, BORDER, CATEGORY_COLORS

logger = logging.getLogger(__name__)


def export_invoices_excel(invoices: list[Invoice], output_path: str | Path) -> Path:
    """Export invoices to a formatted Excel workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    # --- Summary sheet ---
    ws = wb.create_sheet("Resumo Faturas")
    ws.merge_cells("A1:I1")
    ws["A1"] = "Resumo de Faturas"
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["ID", "Fornecedor", "NIF", "N. Fatura", "Data", "Subtotal (EUR)", "IVA (EUR)", "Total (EUR)", "Categoria"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for i, inv in enumerate(invoices, 4):
        cat_display = get_category_display(inv.category) if inv.category else ""
        row_data = [
            inv.id, inv.vendor_name, inv.vendor_nif, inv.invoice_number,
            inv.invoice_date, inv.subtotal, inv.total_iva, inv.total, cat_display,
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.border = BORDER
            if col in (6, 7, 8):
                cell.number_format = "#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            if inv.category and inv.category in CATEGORY_COLORS:
                cell.fill = PatternFill(
                    start_color=CATEGORY_COLORS[inv.category],
                    end_color=CATEGORY_COLORS[inv.category], fill_type="solid",
                )

    # Totals row
    total_row = len(invoices) + 4
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
    for col_idx, col_num in [(6, 6), (7, 7), (8, 8)]:
        cell = ws.cell(row=total_row, column=col_num,
                       value=sum(getattr(inv, ["subtotal", "total_iva", "total"][col_idx - 6]) for inv in invoices))
        cell.font = Font(bold=True)
        cell.number_format = "#,##0.00"
        cell.border = BORDER

    for col in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(3, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 40)

    # --- IVA Summary sheet ---
    ws_iva = wb.create_sheet("Resumo IVA")
    ws_iva.merge_cells("A1:D1")
    ws_iva["A1"] = "Resumo IVA"
    ws_iva["A1"].font = Font(bold=True, size=14)

    iva_headers = ["Taxa IVA (%)", "Base Tributavel (EUR)", "IVA (EUR)", "N. Faturas"]
    for col, h in enumerate(iva_headers, 1):
        cell = ws_iva.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER

    # Aggregate IVA across all invoices
    iva_totals: dict[float, dict] = {}
    for inv in invoices:
        for b in inv.iva_breakdown:
            if b.rate not in iva_totals:
                iva_totals[b.rate] = {"base": 0.0, "iva": 0.0, "count": 0}
            iva_totals[b.rate]["base"] += b.base_amount
            iva_totals[b.rate]["iva"] += b.iva_amount
            iva_totals[b.rate]["count"] += 1

    row = 4
    for rate in sorted(iva_totals.keys()):
        t = iva_totals[rate]
        ws_iva.cell(row=row, column=1, value=f"{rate}%").border = BORDER
        ws_iva.cell(row=row, column=2, value=t["base"]).border = BORDER
        ws_iva.cell(row=row, column=2).number_format = "#,##0.00"
        ws_iva.cell(row=row, column=3, value=t["iva"]).border = BORDER
        ws_iva.cell(row=row, column=3).number_format = "#,##0.00"
        ws_iva.cell(row=row, column=4, value=t["count"]).border = BORDER
        row += 1

    for col in range(1, 5):
        ws_iva.column_dimensions[get_column_letter(col)].width = 22

    # --- Line Items detail sheet ---
    ws_detail = wb.create_sheet("Detalhe Itens")
    ws_detail.merge_cells("A1:H1")
    ws_detail["A1"] = "Detalhe de Itens"
    ws_detail["A1"].font = Font(bold=True, size=14)

    detail_headers = ["Fornecedor", "N. Fatura", "Descricao", "Qtd", "P. Unit (EUR)", "IVA (%)", "IVA (EUR)", "Total (EUR)"]
    for col, h in enumerate(detail_headers, 1):
        cell = ws_detail.cell(row=3, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER

    row = 4
    for inv in invoices:
        for item in inv.line_items:
            row_data = [
                inv.vendor_name, inv.invoice_number, item.description,
                item.quantity, item.unit_price, item.iva_rate, item.iva_amount, item.total,
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws_detail.cell(row=row, column=col, value=val)
                cell.border = BORDER
                if col in (4, 5, 6, 7, 8):
                    cell.number_format = "#,##0.00"
            row += 1

    for col in range(1, len(detail_headers) + 1):
        max_len = max(
            (len(str(ws_detail.cell(row=r, column=col).value or "")) for r in range(3, ws_detail.max_row + 1)),
            default=10,
        )
        ws_detail.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 45)

    wb.save(str(output_path))
    logger.info(f"Invoice Excel exported to {output_path}")
    return output_path


def export_invoices_csv(invoices: list[Invoice], output_path: str | Path) -> Path:
    """Export invoices to CSV."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Fornecedor", "NIF", "N. Fatura", "Data", "Subtotal", "IVA", "Total", "Categoria", "Nota"])
        for inv in invoices:
            cat = get_category_display(inv.category) if inv.category else ""
            writer.writerow([
                inv.vendor_name, inv.vendor_nif, inv.invoice_number,
                inv.invoice_date,
                f"{inv.subtotal:.2f}".replace(".", ","),
                f"{inv.total_iva:.2f}".replace(".", ","),
                f"{inv.total:.2f}".replace(".", ","),
                cat, inv.note or "",
            ])

    logger.info(f"Invoice CSV exported to {output_path}")
    return output_path
