"""Export service for generating Excel and CSV files."""

from __future__ import annotations

import csv
import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from bot.accounting.models import Transaction, ReconciliationResult
from bot.accounting.categorizer import get_category_display

logger = logging.getLogger(__name__)

CATEGORY_COLORS = {
    "estacionamento": "FFF2CC", "transportes": "D9EAD3", "combustivel": "FCE5CD",
    "viagens": "D0E0F0", "alimentacao": "EAD1DC", "receita_vendas": "B6D7A8",
    "receita_servicos": "A4C2F4", "software": "D5A6BD", "material_escritorio": "FFD966",
    "servicos_profissionais": "B4A7D6", "marketing": "FF9900", "fornecedor": "E6B8AF",
    "salarios": "C9DAF8", "aluguer": "D9D2E9", "seguros": "F4CCCC",
    "impostos": "EA9999", "telecomunicacoes": "B7E1CD", "saude": "A2D9CE",
    "transferencia": "D5D5D5", "suprimentos": "CFE2F3", "outros": "EEEEEE",
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _write_transaction_sheet(wb: Workbook, sheet_name: str, transactions: list[Transaction], title: str):
    ws = wb.create_sheet(title=sheet_name)

    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)

    headers = ["Data", "Descricao", "Valor (EUR)", "Tipo", "Categoria", "Nota", "Confianca"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER

    for i, txn in enumerate(transactions, 4):
        cat_display = get_category_display(txn.category) if txn.category else ""
        confidence_map = {"rule": "Auto (regra)", "ai": "IA (sugestao)", "user": "Manual", "unknown": "Nao categorizado"}
        row_data = [
            txn.date, txn.description, txn.value,
            "Debito" if txn.type == "debit" else "Credito",
            cat_display, txn.note or "", confidence_map.get(txn.confidence, txn.confidence),
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=col, value=value)
            cell.border = BORDER
            if col == 3:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right")
            if txn.category and txn.category in CATEGORY_COLORS:
                cell.fill = PatternFill(
                    start_color=CATEGORY_COLORS[txn.category],
                    end_color=CATEGORY_COLORS[txn.category], fill_type="solid",
                )

    for col in range(1, len(headers) + 1):
        max_len = max(
            (len(str(ws.cell(row=r, column=col).value or "")) for r in range(3, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[get_column_letter(col)].width = min(max(max_len + 2, 12), 50)


def _write_summary_sheet(wb: Workbook, result: ReconciliationResult):
    ws = wb.create_sheet(title="Resumo")

    ws.merge_cells("A1:D1")
    ws["A1"] = "Resumo por Categoria"
    ws["A1"].font = Font(bold=True, size=14)

    ws["A3"] = "Ficheiro:"
    ws["B3"] = result.filename
    ws["A4"] = "Data:"
    ws["B4"] = result.parsed_at.strftime("%Y-%m-%d %H:%M")
    ws["A5"] = "Total transacoes:"
    ws["B5"] = len(result.all_transactions)

    if result.bank_balance is not None:
        ws["A6"] = "Saldo bancario:"
        ws["B6"] = result.bank_balance
        ws["B6"].number_format = '#,##0.00'

    headers = ["Categoria", "Debitos (EUR)", "Creditos (EUR)", "N. Transacoes"]
    start_row = 9
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER

    totals: dict[str, dict] = {}
    for txn in result.all_transactions:
        cat = txn.category or "sem_categoria"
        if cat not in totals:
            totals[cat] = {"debits": 0.0, "credits": 0.0, "count": 0}
        if txn.type == "debit":
            totals[cat]["debits"] += txn.value
        else:
            totals[cat]["credits"] += txn.value
        totals[cat]["count"] += 1

    row = start_row + 1
    for cat, t in sorted(totals.items()):
        display = get_category_display(cat) if cat != "sem_categoria" else "Sem Categoria"
        ws.cell(row=row, column=1, value=display).border = BORDER
        ws.cell(row=row, column=2, value=t["debits"]).border = BORDER
        ws.cell(row=row, column=2).number_format = '#,##0.00'
        ws.cell(row=row, column=3, value=t["credits"]).border = BORDER
        ws.cell(row=row, column=3).number_format = '#,##0.00'
        ws.cell(row=row, column=4, value=t["count"]).border = BORDER
        if cat in CATEGORY_COLORS:
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = PatternFill(
                    start_color=CATEGORY_COLORS[cat], end_color=CATEGORY_COLORS[cat], fill_type="solid",
                )
        row += 1

    for col in range(1, 5):
        ws.column_dimensions[get_column_letter(col)].width = 22


def export_excel(result: ReconciliationResult, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)

    _write_summary_sheet(wb, result)
    if result.debit_transactions:
        _write_transaction_sheet(wb, "Debitos Banco", result.debit_transactions, "Movimentos a Debito no Banco")
    if result.credit_transactions:
        _write_transaction_sheet(wb, "Creditos Banco", result.credit_transactions, "Movimentos a Credito no Banco")
    if result.company_debits:
        _write_transaction_sheet(wb, "Debitos Empresa", result.company_debits, "Movimentos a Debito pela Empresa")
    if result.company_credits:
        _write_transaction_sheet(wb, "Creditos Empresa", result.company_credits, "Movimentos a Credito pela Empresa")

    wb.save(str(output_path))
    logger.info(f"Excel exported to {output_path}")
    return output_path


def export_csv(result: ReconciliationResult, output_path: str | Path) -> Path:
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Data", "Descricao", "Valor", "Tipo", "Seccao", "Categoria", "Nota", "Confianca"])
        sections = [
            ("Debitos Banco", result.debit_transactions),
            ("Creditos Banco", result.credit_transactions),
            ("Debitos Empresa", result.company_debits),
            ("Creditos Empresa", result.company_credits),
        ]
        for section_name, transactions in sections:
            for txn in transactions:
                writer.writerow([
                    txn.date, txn.description, f"{txn.value:.2f}".replace(".", ","),
                    "Debito" if txn.type == "debit" else "Credito", section_name,
                    get_category_display(txn.category) if txn.category else "", txn.note or "", txn.confidence,
                ])

    logger.info(f"CSV exported to {output_path}")
    return output_path


def export_pdf(result: ReconciliationResult, output_path: str | Path) -> Path:
    """Export reconciliation as a formatted PDF report."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    doc = SimpleDocTemplate(
        str(output_path), pagesize=A4,
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=20 * mm, bottomMargin=20 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("CustomTitle", parent=styles["Title"], fontSize=16, spaceAfter=12)
    heading_style = ParagraphStyle("CustomHeading", parent=styles["Heading2"], fontSize=12, spaceAfter=6)

    elements = []

    # Title
    elements.append(Paragraph(f"Reconciliacao Bancaria - {result.filename}", title_style))
    elements.append(Paragraph(f"Gerado em: {result.parsed_at.strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
    elements.append(Spacer(1, 10 * mm))

    # Summary table
    summary_data = [
        ["Resumo", ""],
        ["Total transacoes", str(len(result.all_transactions))],
        ["Debitos banco", str(len(result.debit_transactions))],
        ["Creditos banco", str(len(result.credit_transactions))],
        ["Debitos empresa", str(len(result.company_debits))],
        ["Creditos empresa", str(len(result.company_credits))],
    ]
    if result.bank_balance is not None:
        summary_data.append(["Saldo bancario", f"{result.bank_balance:.2f} EUR"])
    if result.reconciled_balance is not None:
        summary_data.append(["Saldo reconciliado", f"{result.reconciled_balance:.2f} EUR"])
    if result.difference is not None:
        summary_data.append(["Diferenca", f"{result.difference:.2f} EUR"])

    summary_table = Table(summary_data, colWidths=[130, 150])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 8 * mm))

    # Transaction sections
    sections = [
        ("Movimentos a Debito no Banco", result.debit_transactions),
        ("Movimentos a Credito no Banco", result.credit_transactions),
        ("Movimentos a Debito pela Empresa", result.company_debits),
        ("Movimentos a Credito pela Empresa", result.company_credits),
    ]

    for section_title, transactions in sections:
        if not transactions:
            continue

        elements.append(Paragraph(section_title, heading_style))

        table_data = [["Data", "Descricao", "Valor (EUR)", "Categoria", "Nota"]]
        for txn in transactions:
            cat_display = get_category_display(txn.category) if txn.category else ""
            desc = txn.description
            if len(desc) > 40:
                desc = desc[:37] + "..."
            note = txn.note or ""
            if len(note) > 30:
                note = note[:27] + "..."
            table_data.append([txn.date, desc, f"{txn.value:.2f}", cat_display, note])

        col_widths = [60, 155, 55, 95, 100]
        t = Table(table_data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F2F2")]),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 6 * mm))

    # Category summary
    elements.append(Paragraph("Resumo por Categoria", heading_style))

    totals = {}
    for txn in result.all_transactions:
        cat = txn.category or "sem_categoria"
        if cat not in totals:
            totals[cat] = {"debits": 0.0, "credits": 0.0, "count": 0}
        if txn.type == "debit":
            totals[cat]["debits"] += txn.value
        else:
            totals[cat]["credits"] += txn.value
        totals[cat]["count"] += 1

    cat_data = [["Categoria", "Debitos (EUR)", "Creditos (EUR)", "N. Transacoes"]]
    for cat, t in sorted(totals.items()):
        display = get_category_display(cat) if cat != "sem_categoria" else "Sem Categoria"
        cat_data.append([display, f"{t['debits']:.2f}", f"{t['credits']:.2f}", str(t["count"])])

    cat_table = Table(cat_data, colWidths=[120, 90, 90, 80])
    cat_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ALIGN", (1, 0), (3, -1), "RIGHT"),
    ]))
    elements.append(cat_table)

    doc.build(elements)
    logger.info(f"PDF exported to {output_path}")
    return output_path
